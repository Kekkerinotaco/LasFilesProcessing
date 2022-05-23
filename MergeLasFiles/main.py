import os
import sort_by_wells
import merge_lases
import openpyxl
from tkinter import *

errors_dict = {}


def main(folder_with_files, result_folders):
    sort_files(folder_with_files, result_folders)
    list_of_objects = os.listdir(result_folders)
    list_of_folders = make_list_of_folders(result_folders, list_of_objects)
    result_folder_path = make_result_folder(result_folders)
    merge_las_files(list_of_folders, result_folder_path)
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    row_counter = 1
    for key, value in errors_dict.items():
        ws["A{}".format(row_counter)] = key
        ws["B{}".format(row_counter)] = str(value)
        row_counter += 1
    wb.save(os.path.join(result_folder_path, "Errors.xlsx"))


def sort_files(folder_with_files, sorted_folders):
    wells_dict = sort_by_wells.get_wells_dict(folder_with_files)
    sort_by_wells.sort_files(sorted_folders, wells_dict, copy=True)


def merge_las_files(list_of_folders, result_folder_path):
    global errors_dict
    for folder in list_of_folders:
        folder_name = os.path.basename(folder)
        result_file_path = os.path.join(result_folder_path, "{}.las".format(folder_name))
        try:
            merge_lases.process_folder(folder, result_file_path)
        except Exception as e:
            if e in errors_dict:
                errors_dict[str(e)].append(folder)
            else:
                errors_dict[str(e)] = []
                errors_dict[str(e)].append(folder)


def make_list_of_folders(sorted_folders, list_of_objects):
    list_of_folders = []
    for object in list_of_objects:
        object_path = os.path.join(sorted_folders, object)
        if os.path.isdir(object_path):
            list_of_folders.append(object_path)
    return list_of_folders


def make_result_folder(sorted_folders):
    result_folder_path = os.path.join(sorted_folders, "!_Result")
    try:
        os.mkdir(result_folder_path)
    except FileExistsError:
        pass
    return result_folder_path

def run_button_click():
    folder_with_files = input_field.get()
    result_folders = export_field.get()
    main(folder_with_files, result_folders)


# Создание рабочего окна
window = Tk()
window.title("Merge las files")
window.geometry("600x250")

# Условные линии сетки
X_FIRST_LANE = 0
X_SECOND_LANE = 350
X_THIRD_LANE = 600

input_label = Label(text="Вставьте ссылку на папку с файлами для обработки:", font=("Arial Bold", 12))
input_label.place(x=X_FIRST_LANE, y=30)
input_field = Entry(width=90)
input_field.place(x=X_FIRST_LANE, y=60)

# Блок лейбл/инпут
export_label = Label(text="Папка, куда сохранятся файлы:", font=("Arial Bold", 12))
export_label.place(x=X_FIRST_LANE, y=110)
export_field = Entry(width=90)
export_field.place(x=X_FIRST_LANE, y=140)

# Кнопка запуска
run_button = Button(text="Провести обработку", command=run_button_click)  # Add a command
run_button.place(x=225, y=175)


# Вызов окна в работу
window.mainloop()

if __name__ == '__main__':
    main()
