import openpyxl
import welly
import os
import time

# Используется для записи значений в результирующем Excel-файле
row_to_write = 2
unique_mnemonics = {}


def main():
    # Ссылка на папку с файлами для обработки
    input_link = r""
    global in_one_cell
    start = time.time()
    result_link = make_result_link(input_link)
    result_workbook = openpyxl.Workbook()
    result_worksheet = result_workbook.worksheets[0]
    write_header(result_worksheet)
    for root, folders, files in os.walk(input_link):
        for file in files:
            if file.endswith(".las"):
                las_path = os.path.join(root, file)
                try:
                    print("Processing file: {}".format(las_path))
                    mnemonics_list = process_las(las_path)
                    write_data(result_worksheet, las_path, mnemonics_list, in_one_cell=in_one_cell)
                except Exception as e:
                    write_error(result_worksheet, las_path, str(e))

    write_unique_mnemonics(result_workbook)
    result_workbook.save(result_link)
    duration = time.time() - start
    print("Duration: {}".format(duration))


def write_header(result_worksheet):
    """ Прописывает заголовок


    :param result_worksheet: Страница Excel, на которой необходимо прописать заголовок
    :return:
    """
    result_worksheet["A1"].value = "Filepath"
    result_worksheet["B1"].value = "Errors"
    result_worksheet["C1"].value = "Mnemonics"


def make_result_link(link):
    """ Формирует ссылку на результирующий файл

    :param link: Ссылка на папку с файлами для обработки
    :return:
    """
    new_link = os.path.join(link, "!_Mnemonics.xlsx")
    return new_link


def process_las(las_path):
    """ Получает Las файл, возвращает список мнемоник в нем, параллельно заполняет счетчик уникальных мнемоник


    :param las_path: Ссылка на las файл
    :return:
    """
    global unique_mnemonics
    well = welly.Well.from_las(las_path, encoding="windows-1251")
    mnemonics_list = [str(item) for item in well.data.keys()]
    for mnemonic in mnemonics_list:
        if mnemonic not in unique_mnemonics.keys():
            unique_mnemonics[mnemonic] = 1
        else:
            unique_mnemonics[mnemonic] += 1
    return mnemonics_list


def write_data(result_worksheet, las_path, mnemonics_list, in_one_cell=True):
    """ Записывает ссылку на файл, и мнемоники содержащиеся в нем в результирующий Excel файл

    :param result_worksheet: Страница Excel для записи значений
    :param las_path: Ссылка на файл
    :param mnemonics_list: Список мнемоник
    :return:
    """
    global row_to_write
    result_worksheet["A{}".format(row_to_write)].value = las_path
    if in_one_cell is True:
        result_worksheet["C{}".format(row_to_write)].value = str(mnemonics_list)
    else:
        for counter in range(1, len(mnemonics_list) + 1):
            result_worksheet.cell(row=row_to_write, column=counter + 2).value = mnemonics_list[counter - 1]
    row_to_write += 1


def write_error(result_worksheet, las_path, error):
    """ Записывает ссылку на файл, и ошибку, случившуюся в нем в результирующий Excel файл

    :param result_worksheet: Страница Excel для записи значений
    :param las_path: Ссылка на файл
    :param error: Ошибка, случившаяся в файле
    :return:
    """
    global row_to_write
    result_worksheet["A{}".format(row_to_write)].value = las_path
    result_worksheet["B{}".format(row_to_write)].value = "{}_occurred_here".format(error)
    row_to_write += 1


def write_unique_mnemonics(result_workbook):
    """ Записывает на странице "Unique_mnemonics" словарь с уникальными мнемониками, и сколько раз они были встречены


    :param result_workbook: Книга Excel, в которой необходимо записать значения
    :return:
    """
    row = 2
    result_worksheet = result_workbook.create_sheet("Unique_mnemonics")
    result_worksheet["A1".format(row)] = "mnemonic"
    result_worksheet["B1".format(row)] = "Files with this mnemonic"
    for mnemonic, number in unique_mnemonics.items():
        result_worksheet["A{}".format(row)] = mnemonic
        result_worksheet["B{}".format(row)] = number
        row += 1


in_one_cell = False
if __name__ == '__main__':
    main()
