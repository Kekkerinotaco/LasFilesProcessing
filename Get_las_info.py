import openpyxl
import lasio
from lasio.exceptions import LASHeaderError
import os

unique_curves = {}

def main():
    # Папка, файлы внутри которой будет необходимо обработать
    link = r""
    result_link = link + "Las_parameters.xlsx"
    result_workbook = openpyxl.Workbook()
    result_worksheet = result_workbook.worksheets[0]
    make_header(result_worksheet)
    for root, folders, files in os.walk(link):
        for file in files:
            if file.endswith(".las"):
                try:
                    file_path = os.path.join(root, file)
                    print(file_path)
                    las = lasio.read(file_path, encoding="windows-1251")
                    well_info = get_well_info(las)
                    write_well_info(result_worksheet, file_path, well_info)
                except:
                    pass
    write_unique_curves(result_worksheet)
    result_workbook.save(result_link)
    print(result_link)


def get_well_info(las):
    """ Функция получает на вход las файл, и получает необходимую информацию из него

    :param las: Las-файл
    :return: Словарь с данными
    """
    global unique_curves
    well_data = {}
    try:
        well_data["Well_name"] = las.well.WELL.value
    except AttributeError:
        well_data["Well_name"] = "Not_present_in_file"

    try:
        well_data["Start_header"] = las.well.STRT.value
    except AttributeError:
        well_data["Start_header"] = "Not_present_in_file"

    las_data = las.df()

    try:
        well_data["Start_MD"] = las_data.index[0]
    except IndexError:
        well_data["Start_MD"] = "Probably_missed_block_delimiter_in_las_file"

    try:
        well_data["Stop_header"] = las.well.STOP.value
    except AttributeError:
        well_data["Stop_header"] = "Not_present_in_file"

    try:
        well_data["Stop_MD"] = las_data.index[-1]
    except IndexError:
        well_data["Stop_MD"] = "Probably_missed_block_delimiter_in_las_file"

    try:
        well_data["Company"] = las.well.COMP.value
    except AttributeError:
        well_data["Company"] = "Not_present_in_file"

    try:
        well_data["Field"] = las.well.FLD.value
    except AttributeError:
        well_data["Field"] = "Not_present_in_file"

    try:
        well_data["Interpretation_company"] = las.well.INTER.value
    except AttributeError:
        well_data["Interpretation_company"] = "Not_present_in_file"

    try:
        well_data["Log_date"] = las.well.SRVC.value
    except AttributeError:
        well_data["Log_date"] = "Not_present_in_file"

    try:
        well_data["Interpretation_date"] = las.well.DATEINTR.value
    except AttributeError:
        well_data["Interpretation_date"] = "Not_present_in_file"

    if well_data["Stop_MD"]:
        try:
            if int(well_data["Stop_MD"]) > 3600:
                well_data["Well_type"] = "H"
            elif int(well_data["Stop_MD"]) < 3600:
                well_data["Well_type"] = "P"
        except ValueError:
            well_data["Well_type"] = "Probably_missed_block_delimiter_in_las_file"

    try:
        get_bitsize(las, well_data)
    except:
        pass

    try:
        well_data["Curves"] = []
    except AttributeError:
        well_data["Curves"] = "Not_present_in_file"

    for curve in las.curves:
        well_data["Curves"].append(curve.mnemonic)
        if curve.mnemonic not in unique_curves:
            unique_curves[curve.mnemonic] = curve.descr
    return well_data


def make_header(worksheet):
    """ Создает шапку таблицы


    :param worksheet: Лист Excel, на котором необходимо прописать значения
    :return:
    """
    worksheet["B2"].value = "File_path"
    worksheet["C2"].value = "Well_name"
    worksheet["D2"].value = "Start_Header"
    worksheet["E2"].value = "Start_MD"
    worksheet["F2"].value = "Stop_Header"
    worksheet["G2"].value = "Stop_MD"
    worksheet["H2"].value = "Company"
    worksheet["I2"].value = "Field"
    worksheet["J2"].value = "Interpretation_company"
    worksheet["K2"].value = "Log_date"
    worksheet["L2"].value = "Interpretation_date"
    worksheet["M2"].value = "Well_type(MD<>3600)"
    worksheet["N2"].value = "BitSize"
    worksheet["O2"].value = "Curves"


def get_bitsize(las, well_data):
    """ Функция достает диаметр долота

   param las: Las файл для обработки
   param well_data: словарь, в котором хранятся значения
    """
    try:
        well_data["Bit_Size"] = las.well.BS.value
    except AttributeError:
        try:
            well_data["Bit_Size"] = las.params.ДНОМ.value
        except:
            well_data["Bit_Size"] = "--"
    except LASHeaderError:
        well_data["Bit_Size"] = "LasheaderError"
    except KeyError:
        well_data["Bit_Size"] = "KeyError"


def write_well_info(worksheet, file_path, well_info):
    """ Прописывает информацию по las файлу


    :param worksheet: Лист Excel, на котором необходимо прописать значения
    :return:
    """
    row_counter = worksheet.max_row + 1
    column = 3
    worksheet["B{}".format(row_counter)].value = file_path
    for dict_value in well_info.values():
        try:
            worksheet.cell(row=row_counter, column=column).value = dict_value
        except:
            worksheet.cell(row=row_counter, column=column).value = str(dict_value)
        column += 1


def write_unique_curves(worksheet):
    """ Прописывает словарь с уникальными мнемониками, и их расшифровками


    :param worksheet: Лист Excel, на котором необходимо прописать значения
    :return:
    """
    global unique_curves
    worksheet["P2"].value = "Curve_mnemonic"
    worksheet["Q2"].value = "Description"
    row_number = 3
    for mnemonic, description in unique_curves.items():
        worksheet["R{}".format(row_number)].value = mnemonic
        worksheet["S{}".format(row_number)].value = description
        row_number += 1


if __name__ == '__main__':
    main()
